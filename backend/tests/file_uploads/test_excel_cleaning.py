"""
Tests for Excel file cleaning functionality in utils_file_uploads module.
"""
import pytest
import pandas as pd
import io
import openpyxl
from openpyxl.styles import Alignment
from utils_file_uploads import ExcelUtils


class TestExcelCleaning:
    """Tests for Excel file cleaning functionality."""

    @pytest.fixture
    def sample_excel_with_merged_cells(self):
        """Create a sample Excel file with merged cells."""
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "SheetWithMergedCells"

        # Add some headers with merged cells
        sheet.merge_cells('A1:C1')
        sheet['A1'] = 'Merged Header'
        sheet['A1'].alignment = Alignment(horizontal='center')

        # Add regular headers
        sheet['A2'] = 'ID'
        sheet['B2'] = 'Name'
        sheet['C2'] = 'Value'

        # Add some data
        data = [
            [1, 'Item 1', 10.5],
            [2, 'Item 2', 20.5],
            [3, 'Item 3', 30.5],
        ]
        for row_idx, row_data in enumerate(data, start=3):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Add another merged cell in the data
        sheet.merge_cells('B5:C5')
        sheet['B5'] = 'Merged Data'
        sheet['B5'].alignment = Alignment(horizontal='center')

        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        return output

    @pytest.fixture
    def sample_excel_with_empty_rows_cols(self):
        """Create a sample Excel file with empty rows and columns."""
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "EmptyRowsCols"

        # Add headers
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Name'
        sheet['D1'] = 'Value'  # Skip column C

        # Add data with empty rows
        data = [
            [1, 'Item 1', None, 10.5],
            None,  # Empty row
            [2, 'Item 2', None, 20.5],
            None,  # Empty row
            [3, 'Item 3', None, 30.5],
        ]
        for row_idx, row_data in enumerate(data, start=2):
            if row_data is not None:
                for col_idx, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        return output

    @pytest.fixture
    def sample_excel_with_multiple_sheets(self):
        """Create a sample Excel file with multiple sheets."""
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        
        # First sheet
        sheet1 = wb.active
        sheet1.title = "Sheet1"
        sheet1['A1'] = 'ID'
        sheet1['B1'] = 'Name'
        sheet1['C1'] = 'Value'
        
        data1 = [
            [1, 'Item 1', 10.5],
            [2, 'Item 2', 20.5],
        ]
        for row_idx, row_data in enumerate(data1, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet1.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Second sheet
        sheet2 = wb.create_sheet(title="Sheet2")
        sheet2['A1'] = 'Category'
        sheet2['B1'] = 'Count'
        
        data2 = [
            ['A', 5],
            ['B', 10],
            ['C', 15],
        ]
        for row_idx, row_data in enumerate(data2, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet2.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        return output

    @pytest.mark.asyncio
    async def test_clean_excel_pd_merged_cells(self, sample_excel_with_merged_cells):
        """Test cleaning Excel with merged cells."""
        # Process the excel file
        result = await ExcelUtils.clean_excel_pd(sample_excel_with_merged_cells)
        
        # Check that we got a dictionary with sheet name as key
        assert isinstance(result, dict)
        assert "sheetwithmergedcells" in result
        
        # Get the sheet data
        sheet_name = next(iter(result))
        df = result[sheet_name]
        
        # Check that all cells in the first row have the merged header value
        assert df.iloc[0, 0] == "Merged Header"
        assert df.iloc[0, 1] == "Merged Header"
        assert df.iloc[0, 2] == "Merged Header"
        
        # Check that data rows were properly parsed
        assert df.iloc[1, 0] == "ID"
        assert df.iloc[1, 1] == "Name"
        assert df.iloc[1, 2] == "Value"
        
        assert df.iloc[2, 0] == 1
        assert df.iloc[2, 1] == "Item 1"
        assert df.iloc[2, 2] == 10.5
        
        # Check the merged data cell was forward-filled
        assert df.iloc[4, 1] == "Merged Data"
        assert df.iloc[4, 2] == "Merged Data"

    @pytest.mark.asyncio
    async def test_clean_excel_pd_empty_rows_cols(self, sample_excel_with_empty_rows_cols):
        """Test cleaning Excel with empty rows and columns."""
        # Process the excel file
        result = await ExcelUtils.clean_excel_pd(sample_excel_with_empty_rows_cols)
        
        # Get the sheet data
        sheet_name = next(iter(result))
        df = result[sheet_name]
        
        # Check that empty rows were removed
        # Original had 5 rows plus header, but with 2 empty rows
        # So we should have 4 rows total after cleaning
        assert len(df) == 4
        
        # Check that columns were compressed (column C was empty)
        # Expecting 3 columns (ID, Name, Value)
        assert df.shape[1] == 3
        
        # Check data integrity
        assert df.iloc[0, 0] == "ID"
        assert df.iloc[0, 1] == "Name"
        assert df.iloc[0, 2] == "Value"
        
        # Data rows should be continuous after empty row removal
        assert df.iloc[1, 0] == 1
        assert df.iloc[2, 0] == 2
        assert df.iloc[3, 0] == 3

    @pytest.mark.asyncio
    async def test_clean_excel_pd_multiple_sheets(self, sample_excel_with_multiple_sheets):
        """Test cleaning Excel with multiple sheets."""
        # Process the excel file
        result = await ExcelUtils.clean_excel_pd(sample_excel_with_multiple_sheets)
        
        # Check that we got a dictionary with both sheet names as keys
        assert isinstance(result, dict)
        assert len(result) == 2
        
        # Check first sheet
        assert "sheet1" in result
        df1 = result["sheet1"]
        
        assert df1.shape == (3, 3)  # 2 data rows + 1 header row, 3 columns
        assert df1.iloc[0, 0] == "ID"
        assert df1.iloc[1, 0] == 1
        assert df1.iloc[2, 0] == 2
        
        # Check second sheet
        assert "sheet2" in result
        df2 = result["sheet2"]
        
        assert df2.shape == (4, 2)  # 3 data rows + 1 header row, 2 columns
        assert df2.iloc[0, 0] == "Category"
        assert df2.iloc[1, 0] == "A"
        assert df2.iloc[2, 0] == "B"
        assert df2.iloc[3, 0] == "C"

    @pytest.mark.asyncio
    async def test_clean_excel_pd_empty_sheet(self):
        """Test cleaning Excel with empty sheet."""
        # Create an empty Excel file
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "EmptySheet"
        wb.save(output)
        output.seek(0)
        
        # Process the excel file
        result = await ExcelUtils.clean_excel_pd(output)
        
        # Check that we got an empty dataframe
        assert isinstance(result, dict)
        assert "emptysheet" in result
        df = result["emptysheet"]
        
        assert df.empty
        assert df.shape[0] == 0  # No rows
        
    @pytest.mark.asyncio
    async def test_is_excel_dirty_clean(self):
        """Test detecting clean Excel data that doesn't need OpenAI cleaning."""
        # Create a simple clean dataframe
        df = pd.DataFrame({
            'ID': [1, 2, 3, 4, 5],
            'Name': ['Product A', 'Product B', 'Product C', 'Product D', 'Product E'],
            'Price': [10.5, 20.75, 15.25, 30.00, 25.50]
        })
        
        # Check if the dataframe is dirty
        is_dirty = await ExcelUtils.is_excel_dirty("clean_table", df)
        
        # Should be clean
        assert not is_dirty
        
    @pytest.mark.asyncio
    async def test_is_excel_dirty_with_headers(self):
        """Test detecting Excel data with repeated headers that need cleaning."""
        # Create a dataframe with repeated headers
        df = pd.DataFrame([
            ['Company Report', 'Company Report', 'Company Report'],
            ['Q1 2023', 'Q1 2023', 'Q1 2023'],
            ['ID', 'Name', 'Price'],
            [1, 'Product A', 10.5],
            [2, 'Product B', 20.75],
            [3, 'Product C', 15.25]
        ])
        
        # Check if the dataframe is dirty
        is_dirty = await ExcelUtils.is_excel_dirty("header_table", df)
        
        # Should be dirty due to repeated headers
        assert is_dirty
        
    @pytest.mark.asyncio
    async def test_is_excel_dirty_with_footers(self):
        """Test detecting Excel data with footer notes that need cleaning."""
        # Create a dataframe with footer notes
        df = pd.DataFrame([
            ['ID', 'Name', 'Price'],
            [1, 'Product A', 10.5],
            [2, 'Product B', 20.75],
            [3, 'Product C', 15.25],
            ['Note:', 'All prices are in USD', ''],
            ['Source:', 'Internal data', '']
        ])
        
        # Check if the dataframe is dirty
        is_dirty = await ExcelUtils.is_excel_dirty("footer_table", df)
        
        # Should be dirty due to footer notes
        assert is_dirty
        
    @pytest.mark.asyncio
    async def test_is_excel_dirty_with_totals(self):
        """Test detecting Excel data with aggregate statistics that need cleaning."""
        # Create a dataframe with totals
        df = pd.DataFrame([
            ['ID', 'Name', 'Price'],
            [1, 'Product A', 10.5],
            [2, 'Product B', 20.75],
            [3, 'Product C', 15.25],
            ['Total', '', 46.5]
        ])
        
        # Check if the dataframe is dirty
        is_dirty = await ExcelUtils.is_excel_dirty("totals_table", df)
        
        # Should be dirty due to totals
        assert is_dirty
        
    @pytest.mark.asyncio
    async def test_is_excel_dirty_wide_format(self):
        """Test detecting Excel data in wide format that needs cleaning."""
        # Create a dataframe in wide format
        df = pd.DataFrame({
            'Product': ['Product A', 'Product B', 'Product C'],
            'Q1 Sales': [100, 200, 300],
            'Q2 Sales': [150, 250, 350],
            'Q3 Sales': [120, 220, 320],
            'Q4 Sales': [180, 280, 380],
            'Q1 Profit': [50, 100, 150],
            'Q2 Profit': [75, 125, 175],
            'Q3 Profit': [60, 110, 160],
            'Q4 Profit': [90, 140, 190]
        })
        
        # Check if the dataframe is dirty
        is_dirty = await ExcelUtils.is_excel_dirty("wide_format_table", df)
        
        # Should be dirty due to wide format
        assert is_dirty