import re
from uuid import uuid4
import numpy as np
import pandas as pd
import datetime
from sqlalchemy import text
from sqlalchemy.ext.asyncio import create_async_engine
from dateutil import parser
from io import BytesIO
import openpyxl
import asyncio
import os
import concurrent.futures
from openai import AsyncOpenAI
from defog.llm.utils import LLM_COSTS_PER_TOKEN
from utils_logging import LOGGER

# PostgreSQL reserved words for column name sanitization
POSTGRES_RESERVED_WORDS = {
    "select", "from", "where", "join", "table", "order", "group", "by", 
    "create", "drop", "insert", "update", "delete", "alter", "column", "user", 
    "and", "or", "not", "null", "true", "false", "primary", "key", "foreign", 
    "unique", "check", "default", "index", "on", "as", "asc", "desc", "varchar", 
    "int", "bigint", "float", "decimal", "text", "boolean", "date", "timestamp",
}


class NameUtils:
    """Utilities for handling table and column names."""

    @staticmethod
    def clean_table_name(table_name: str, existing=None):
        """
        Cleans a table name by snake casing it and making it lower case.

        Args:
            table_name: Table name to clean
            existing: List of existing table names to avoid duplicates

        Returns:
            Cleaned table name
        """
        if existing is None:
            existing = []

        if not isinstance(table_name, str):
            raise ValueError("Table name must be a string.")

        validated = str(table_name).strip().lower()
        validated = re.sub(r"[^a-zA-Z0-9_]", "_", validated)

        if validated in existing:
            validated = f"{validated}_{uuid4().hex[:7]}"

        if not validated:
            validated = f"table_{uuid4().hex[:7]}"

        return validated

    @staticmethod
    def sanitize_column_name(col_name: str):
        """
        Convert a column name into a "safe" Postgres identifier.

        Args:
            col_name: Column name to sanitize

        Returns:
            Sanitized column name
        """
        if not isinstance(col_name, str):
            col_name = str(col_name)

        col_name = col_name.strip().lower()

        # Replace special characters
        col_name = col_name.replace("%", "perc")
        col_name = col_name.replace("&", "and")

        # Replace invalid characters with underscores
        col_name = re.sub(r"[^a-z0-9_]", "_", col_name)

        # Collapse multiple underscores into single
        col_name = re.sub(r"_+", "_", col_name)

        # Strip leading/trailing underscores
        col_name = col_name.strip("_")

        # If empty after stripping, fallback
        if not col_name:
            col_name = "col"

        # If it starts with digit, prepend underscore
        if re.match(r"^\d", col_name):
            col_name = f"_{col_name}"

        # If it's a reserved keyword, add suffix
        if col_name in POSTGRES_RESERVED_WORDS:
            col_name = f"{col_name}_col"

        return col_name


class DateTimeUtils:
    """Utilities for handling date and time operations."""

    # Common terms that indicate date columns
    DATE_TERMS = [
        "date", "time", "year", "month", "day", "quarter", "qtr", 
        "yr", "mm", "dd", "yyyy", "created", "modified", "updated", 
        "timestamp", "dob", "birth", "start", "end", "period", 
        "calendar", "fiscal", "dtm", "dt", "ymd", "mdy", "dmy",
    ]

    # Common terms that indicate time-only columns
    TIME_TERMS = [
        "time", "hour", "minute", "second", "hr", "min", "sec",
        "am", "pm", "duration", "interval"
    ]

    # Common date patterns for regex matching
    COMMON_DATE_PATTERNS = [
        # MM/DD/YYYY or DD/MM/YYYY
        r"^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$",
        # YYYY/MM/DD
        r"^\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2}$",
        # Month name formats: Jan 01, 2020 or January 1, 2020
        r"^[A-Za-z]{3,9}\.?\s+\d{1,2},?\s+\d{2,4}$",
        # 01-Jan-2020 or 1-January-20
        r"^\d{1,2}[/\-\.\s]+[A-Za-z]{3,9}\.?[/\-\.\s]+\d{2,4}$",
        # Formats like 01Jan2023 or 01JAN2023
        r"^\d{1,2}[A-Za-z]{3,9}\d{2,4}$",
        # Short date format like MM/DD or MM-DD
        r"^\d{1,2}[/\-\.]\d{1,2}$",
    ]

    # Common time patterns for regex matching
    TIME_PATTERNS = [
        # HH:MM or HH:MM:SS (24-hour format)
        r"^([01]\d|2[0-3]):[0-5]\d(:[0-5]\d)?$",
        # H:MM or H:MM:SS (single digit hour)
        r"^[0-9]:[0-5]\d(:[0-5]\d)?$",
        # HH:MM AM/PM or HH:MM:SS AM/PM (12-hour format)
        r"^(0?[1-9]|1[0-2]):[0-5]\d(:[0-5]\d)?\s*[AaPp][Mm]$",
        # Military time (0000-2359)
        r"^([01]\d|2[0-3])([0-5]\d)$",
    ]

    @classmethod
    def is_date_column_name(cls, col_name):
        """
        Check if a column name indicates it might contain date/time data.

        Args:
            col_name: Column name to check

        Returns:
            True if the column name suggests date content
        """
        if not isinstance(col_name, str):
            return False

        # Normalize the column name for better matching
        name_lower = col_name.lower()

        # Check for ID patterns - don't treat as dates
        if (
            name_lower == "id"
            or name_lower.endswith("_id")
            or name_lower.startswith("id_")
            or "_id_" in name_lower
        ):
            return False

        # Check if any date term appears in the column name
        for term in cls.DATE_TERMS:
            if term in name_lower:
                return True

        # Check for common date patterns
        if re.search(r"(_|^)(dt|date|time)(_|$)", name_lower):
            return True

        return False

    @classmethod
    def is_time_column_name(cls, col_name):
        """
        Check if a column name indicates it might contain time-only data.

        Args:
            col_name: Column name to check

        Returns:
            True if the column name suggests time content
        """
        if not isinstance(col_name, str):
            return False

        # Normalize the column name for better matching
        name_lower = col_name.lower()

        # Check for ID patterns
        if (
            name_lower == "id"
            or name_lower.endswith("_id")
            or name_lower.startswith("id_")
            or "_id_" in name_lower
        ):
            return False

        # Simple direct check for common time column patterns
        if (
            "_time" in name_lower
            or name_lower.startswith("time_")
            or name_lower == "time"
            or name_lower in ["hour", "minute", "second"]
        ):

            # Exclude patterns that suggest datetime rather than just time
            if not any(
                date_term in name_lower
                for date_term in ["date", "timestamp", "datetime"]
            ):
                return True

        # Date terms that would indicate this is a full date/timestamp, not just time
        for term in cls.DATE_TERMS:
            if term in name_lower and term not in cls.TIME_TERMS:
                return False

        # Check if any time-specific term appears in the name
        for term in cls.TIME_TERMS:
            # Match whole words or components with word boundaries
            if (
                re.search(rf"(^|_){term}($|_)", name_lower)
                or name_lower == term
                or name_lower.endswith(f"_{term}")
                or name_lower.startswith(f"{term}_")
            ):
                return True

        return False

    @classmethod
    def can_parse_date(cls, val):
        """
        Check if a value can be parsed as a date.

        Args:
            val: Value to check

        Returns:
            True if the value appears to be a date
        """
        if not isinstance(val, str):
            val = str(val)

        val = val.strip()

        # Quick rejections
        if (
            not val
            or val.lower() in ("invalid date", "not a date", "na", "n/a")
            or len(val) <= 4
        ):
            return False

        # Detect date ranges like "Feb 14-21" or "Feb 26 - Mar 4" and exclude them
        if re.search(
            r"([A-Za-z]{3,9}\.?\s+\d{1,2}\s*[-–]\s*\d{1,2})", val
        ) or re.search(r"([A-Za-z]{3,9}\.?\s+\d{1,2}\s*[-–]\s*[A-Za-z]{3,9})", val):
            return False

        # Check against common date patterns
        for pattern in cls.COMMON_DATE_PATTERNS:
            if re.match(pattern, val):
                try:
                    parser.parse(val)
                    return True
                except:
                    pass

        # Handle all-digit strings
        if re.fullmatch(r"\d+", val):
            if len(val) in (6, 8):  # YYMMDD or YYYYMMDD format
                try:
                    parser.parse(val)
                    return True
                except Exception:
                    return False
            return False

        # For strings without date separators
        if not re.search(r"[\s/\-\.:]", val):
            try:
                parsed_date = parser.parse(val, fuzzy=False)
                return 1900 <= parsed_date.year <= 2100
            except:
                return False

        # Special handling for US-style dates
        if re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", val):
            try:
                parsed_date = parser.parse(val, fuzzy=False)
                return 1900 <= parsed_date.year <= 2100
            except:
                pass

        # Final attempt with dateutil parser
        try:
            parsed_date = parser.parse(val, fuzzy=False)
            return 1900 <= parsed_date.year <= 2100
        except Exception:
            return False

    @classmethod
    def can_parse_time(cls, val):
        """
        Check if a value can be parsed as a time-only value.

        Args:
            val: Value to check

        Returns:
            True if the value appears to be a time
        """
        if not isinstance(val, str):
            val = str(val)

        val = val.strip()

        # Quick rejections
        if not val or len(val) < 3 or re.match(r"^(19|20)\d{2}$", val):
            return False

        # Check against time patterns
        for pattern in cls.TIME_PATTERNS:
            if re.match(pattern, val):
                try:
                    # Special handling for military time
                    if len(val) == 4 and val.isdigit():
                        return bool(re.match(r"^([01]\d|2[0-3])([0-5]\d)$", val))
                    _ = parser.parse(val).time()
                    return True
                except Exception:
                    pass

        # Try more general time parsing
        try:
            parsed = parser.parse(val)

            # Check if this looks like a time-only value:
            # 1. Has colon (common in time)
            # 2. Doesn't have date separators
            if ":" in val and not any(x in val for x in ["/", "-", ".", ","]):
                # Check if the original string contains the time component
                # but not common month/day names
                time_str = parsed.strftime("%H:%M")
                month_day_terms = [
                    "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", 
                    "oct", "nov", "dec", "monday", "tuesday", "wednesday", "thursday",
                    "friday", "saturday", "sunday"
                ]
                
                if time_str in val and not any(x in val.lower() for x in month_day_terms):
                    return True

            return False
        except Exception:
            return False


class TypeUtils:
    """Utilities for type detection and conversion."""

    @staticmethod
    def to_float_if_possible(val):
        """
        Attempt to parse a value to numeric after cleaning.

        Args:
            val: Value to convert

        Returns:
            Float value if conversion is possible, None otherwise
        """
        if not val:
            return None

        try:
            # Handle accounting negative numbers (123.45) -> -123.45
            val_str = str(val).strip()
            if val_str.startswith("(") and val_str.endswith(")"):
                val_str = "-" + val_str[1:-1].strip("$").strip()

            # Clean the value by removing non-numeric symbols
            cleaned_val = re.sub(r"[^\d.\-+eE]", "", val_str)

            # Basic validations
            if cleaned_val in ("", ".", "-", "+"):
                return None

            digit_count = sum(c.isdigit() for c in val_str)
            alphanum_count = sum(c.isalnum() for c in val_str)

            if digit_count == 0 or alphanum_count == 0 or alphanum_count > digit_count:
                return None

            # Try parsing as float
            return float(cleaned_val)
        except:
            return None

    @staticmethod
    def guess_column_type(series, column_name=None, sample_size=50):
        """
        Guess the most appropriate Postgres column type for a data series.

        Args:
            series: Pandas Series to analyze
            column_name: Original column name for heuristics
            sample_size: Number of samples to take for analysis

        Returns:
            PostgreSQL data type as string
        """
        # Drop nulls/empty values
        non_null_values = [str(v) for v in series.dropna() if str(v).strip() != ""]

        # If there's nothing in this column, assume TEXT
        if len(non_null_values) == 0:
            return "TEXT"

        # Sample values to limit computational overhead
        sampled_values = (
            non_null_values[:sample_size]
            if len(non_null_values) > sample_size
            else non_null_values
        )

        # Check column name heuristics
        column_suggests_date = (
            column_name is not None and DateTimeUtils.is_date_column_name(column_name)
        )
        column_suggests_time = (
            column_name is not None and DateTimeUtils.is_time_column_name(column_name)
        )

        # Check if column is ID or suggests numeric values
        column_is_id = False
        column_suggests_numeric = False

        if column_name is not None:
            column_name = str(column_name)
            name_lower = column_name.lower()
            # Check for ID patterns
            if (
                name_lower == "id"
                or name_lower.endswith("_id")
                or name_lower.startswith("id_")
                or "_id_" in name_lower
            ):
                column_is_id = True
                column_suggests_date = False
                column_suggests_time = False
                column_suggests_numeric = True

            # Check for numeric indicator terms
            numeric_terms = [
                "amount", "total", "sum", "value", "price", "cost", "revenue", "income",
                "expense", "fee", "rate", "ratio", "score", "balance", "million", "billion",
                "thousand", "usd", "eur", "gbp", "jpy", "dollar", "euro", "pound", "yen",
                "currency", "money", "cash", "profit", "loss", "gain", "discount", "tax",
                "interest", "count", "number", "quantity", "weight", "height", "width",
                "length", "volume", "area", "size", "measurement", "percentage", "percent"
            ]

            col_lower = name_lower.replace("_", " ")
            for term in numeric_terms:
                if term in col_lower:
                    column_suggests_numeric = True
                    break

        # Check for obvious text values
        has_obvious_text = any(
            re.search(r"[a-df-zA-DF-Z]", v)
            and not DateTimeUtils.can_parse_date(v)
            and not DateTimeUtils.can_parse_time(v)
            for v in sampled_values
        )

        # Count different value types
        pct_count = sum(1 for v in sampled_values if str(v).strip().endswith("%"))
        time_count = sum(1 for v in sampled_values if DateTimeUtils.can_parse_time(v))
        date_count = sum(1 for v in sampled_values if DateTimeUtils.can_parse_date(v))
        us_date_count = sum(
            1
            for v in sampled_values
            if re.match(r"^(\d{1,2}|\d{4})/\d{1,2}/(\d{2}|\d{4})$", v.strip())
        )
        decimal_count = sum(
            1 for v in sampled_values if re.match(r"^-?\$?[0-9,]+\.\d+$", v.strip())
        )
        short_date_count = sum(
            1 for v in sampled_values if re.match(r"^\d{1,2}[/\-\.]\d{1,2}$", v.strip())
        )
        sci_notation_count = sum(
            1
            for v in sampled_values
            if re.search(r"^-?\d*\.?\d+[eE][+-]?\d+$", v.strip())
        )
        yyyymmdd_count = sum(
            1
            for v in sampled_values
            if re.fullmatch(r"\d{8}", v) and DateTimeUtils.can_parse_date(v)
        )
        iso_date_count = sum(
            1 for v in sampled_values if re.match(r"^\d{4}-\d{2}-\d{2}", v.strip())
        )

        # Calculate ratios
        total_samples = len(sampled_values)
        pct_ratio = pct_count / total_samples
        time_ratio = time_count / total_samples
        date_ratio = min(date_count / total_samples, 1.0)
        us_date_ratio = us_date_count / total_samples
        decimal_ratio = decimal_count / total_samples
        short_date_ratio = short_date_count / total_samples
        sci_notation_ratio = sci_notation_count / total_samples
        yyyymmdd_ratio = yyyymmdd_count / total_samples
        iso_date_ratio = iso_date_count / total_samples

        # Parse numeric values
        float_parsed = [TypeUtils.to_float_if_possible(v) for v in sampled_values]
        numeric_count = sum(x is not None for x in float_parsed)
        numeric_ratio = numeric_count / total_samples

        # Decision tree for type inference

        # Percentage check
        if pct_ratio > 0.3:
            return "DOUBLE PRECISION" if pct_ratio > 0.8 else "TEXT"

        # Time check
        if time_ratio > 0.7 or (column_suggests_time and time_ratio > 0.5):
            return "TIME"

        # US-style date check
        if us_date_ratio > 0.7:
            return "TIMESTAMP"

        # Decimal check
        if decimal_ratio > 0.7:
            return "DOUBLE PRECISION"

        # Short date check
        if short_date_ratio > 0.7:
            return "TIMESTAMP"

        # Scientific notation check
        if sci_notation_ratio > 0.2:
            return "DOUBLE PRECISION"

        # YYYYMMDD format date check
        if yyyymmdd_ratio > 0.7:
            return "TIMESTAMP"

        # ISO8601 date check
        if iso_date_ratio > 0.7:
            return "TIMESTAMP"

        # Check for text values in full dataset
        has_any_text_in_full = False
        text_value_count = 0
        for value in non_null_values:
            if (
                re.search(r"[a-zA-Z]", value)
                and not re.search(r"[eE][-+]?\d+", value)
                and not DateTimeUtils.can_parse_date(value)
                and not DateTimeUtils.can_parse_time(value)
            ):
                has_any_text_in_full = True
                text_value_count += 1

        # If significant text values and not date/time column, use TEXT
        if (
            has_any_text_in_full
            and not (column_suggests_date or column_suggests_time)
            and (text_value_count / len(non_null_values)) > 0.05
        ):
            return "TEXT"

        # If obvious text values, prefer TEXT
        if (has_obvious_text or has_any_text_in_full) and not (
            numeric_ratio > 0.95
            or date_ratio > 0.95
            or time_ratio > 0.95
            or column_suggests_date
            or column_suggests_time
        ):
            return "TEXT"

        # Special case for year columns
        if (
            column_name
            and (
                column_name.lower() == "year"
                or re.match(r"^(fiscal_|calendar_)?years?(_\d+)?$", column_name.lower())
                or "year" in column_name.lower()
            )
            and numeric_ratio > 0.8
        ):

            # Check if values appear to be years
            are_years = all(
                val is None or (val.is_integer() and 1000 <= val <= 2100)
                for val in float_parsed
                if val is not None
            )

            if are_years and any(float_parsed):
                return "BIGINT"

        # Time columns with confirmed time values
        if column_suggests_time and time_ratio > 0.7:
            return "TIME"

        # Date columns with confirmed date values
        if column_suggests_date and date_ratio > 0.7:
            # Check for month name columns
            month_names = [
                "jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
                "january", "february", "march", "april", "may", "june", "july", "august",
                "september", "october", "november", "december"
            ]

            month_name_count = sum(
                1 for val in sampled_values if str(val).lower().strip() in month_names
            )

            month_ratio = month_name_count / total_samples

            # If only month names, they should be TEXT
            if month_ratio > 0.7 and month_name_count == total_samples:
                return "TEXT"

            # Check for "invalid date" text
            invalid_text_ratio = (
                sum(
                    1
                    for v in sampled_values
                    if re.search(r"invalid|not\s+a\s+date", str(v).lower())
                )
                / total_samples
            )

            if invalid_text_ratio > 0.25 and column_name != "created_date":
                return "TEXT"

            # Otherwise, use TIMESTAMP
            return "TIMESTAMP"

        # Check if enough values are numeric
        numeric_threshold = 0.7 if column_suggests_numeric else 0.8
        if numeric_ratio >= numeric_threshold:
            # Check if all numeric values are integers
            are_ints = [val is not None and val.is_integer() for val in float_parsed]

            if all(are_ints) and are_ints:
                # Date or time column check with numeric values
                if column_suggests_date and "year" not in str(column_name).lower():
                    if (
                        sum(
                            1 for v in sampled_values if DateTimeUtils.can_parse_date(v)
                        )
                        / total_samples
                        > 0.6
                    ):
                        return "TIMESTAMP"

                if column_suggests_time:
                    if (
                        sum(
                            1 for v in sampled_values if DateTimeUtils.can_parse_time(v)
                        )
                        / total_samples
                        > 0.6
                    ):
                        return "TIME"

                # Short date patterns check
                if short_date_count > 0 and short_date_ratio > 0.5:
                    return "TIMESTAMP"

                # Final text check
                if any(
                    re.search(r"[a-zA-Z]", val) and not re.search(r"[eE][-+]?\d+", val)
                    for val in non_null_values
                ):
                    return "TEXT"

                return "BIGINT"
            else:
                return "DOUBLE PRECISION"

        # Lower threshold checks
        if time_ratio > 0.6:
            return "TIME"

        if date_ratio > 0.7:
            return "TIMESTAMP"

        # Default to TEXT
        return "TEXT"

    @staticmethod
    def convert_values_to_postgres_type(value, target_type: str):
        """
        Convert a value to the appropriate Python object for insertion into Postgres.

        Args:
            value: Value to convert
            target_type: PostgreSQL type string

        Returns:
            Converted value appropriate for the target type, or None if conversion fails
        """
        # Handle Pandas Series objects
        if isinstance(value, pd.Series):
            return value.apply(
                lambda x: TypeUtils.convert_values_to_postgres_type(x, target_type)
            )

        # Handle None and NaN values
        if (
            value is None
            or (isinstance(value, float) and pd.isna(value))
            or pd.isna(value)
        ):
            return None

        val_str = str(value).strip()

        # Handle common NULL-like string values
        if val_str.lower() in ("", "null", "none", "nan", "   "):
            return None

        if target_type == "TIME":
            if not DateTimeUtils.can_parse_time(val_str):
                return None

            # Parse military time format directly
            if re.match(r"^([01]\d|2[0-3])([0-5]\d)$", val_str):
                hour = int(val_str[:2])
                minute = int(val_str[2:])
                return datetime.time(hour, minute)

            try:
                # Try various time parsing approaches
                parsed_time = parser.parse(val_str).time()
                return parsed_time
            except Exception:
                try:
                    # HH:MM(:SS) format
                    if re.match(
                        r"^([01]?\d|2[0-3]):([0-5]\d)(?::([0-5]\d))?$", val_str
                    ):
                        parts = val_str.split(":")
                        hour = int(parts[0])
                        minute = int(parts[1])
                        second = int(parts[2]) if len(parts) > 2 else 0
                        return datetime.time(hour, minute, second)

                    # AM/PM times
                    if re.match(
                        r"^([0-9]|1[0-2]):([0-5][0-9])(?::([0-5][0-9]))?\s*([AaPp][Mm])$",
                        val_str,
                    ):
                        is_pm = val_str.lower().endswith("pm")
                        time_part = val_str[:-2].strip()
                        parts = time_part.split(":")
                        hour = int(parts[0])
                        minute = int(parts[1])
                        second = int(parts[2]) if len(parts) > 2 else 0

                        if is_pm and hour < 12:
                            hour += 12
                        elif not is_pm and hour == 12:
                            hour = 0

                        return datetime.time(hour, minute, second)
                except:
                    pass
                return None

        elif target_type == "TIMESTAMP":
            # Don't parse scientific notation as dates
            if re.search(r"^-?\d*\.?\d+[eE][+-]?\d+$", val_str.strip()):
                return None

            # Check if it can be parsed as a date
            if not DateTimeUtils.can_parse_date(val_str):
                return None

            # Check for invalid date patterns
            if re.search(r"(\d{4}-\d{2}-\d{2})-[a-zA-Z]", val_str):
                return None

            try:
                parsed_date = parser.parse(val_str)
                # Verify the year is reasonable
                if 1900 <= parsed_date.year <= 2100:
                    return parsed_date
                return None
            except Exception:
                return None

        elif target_type in ("BIGINT", "DOUBLE PRECISION"):
            # Handle accounting negative numbers (123.45) -> -123.45
            if val_str.startswith("(") and val_str.endswith(")"):
                val_str = "-" + val_str[1:-1].strip("$").strip()

            # Handle percentage values
            if val_str.endswith("%"):
                val_str = val_str.rstrip("%").strip()
                try:
                    return float(re.sub(r"[^\d.\-+eE]", "", val_str)) / 100
                except:
                    return None

            # Handle currency codes
            if re.search(r"\s+[A-Za-z]{3}$", val_str):
                val_str = re.sub(r"\s+[A-Za-z]{3}$", "", val_str)
            elif re.search(r"^[A-Za-z]{3}\s+", val_str):
                val_str = re.sub(r"^[A-Za-z]{3}\s+", "", val_str)

            # Skip processing for values with letters (except scientific notation)
            if re.search(r"[a-zA-Z]", val_str) and not re.search(
                r"[eE][-+]?\d+", val_str
            ):
                return None

            # Scientific notation handling
            if re.search(r"^-?\d*\.?\d+[eE][+-]?\d+$", val_str.strip()):
                try:
                    float_val = float(val_str)
                    if target_type == "BIGINT":
                        if pd.isna(float_val) or float_val in (
                            float("inf"),
                            float("-inf"),
                        ):
                            return None
                        # Check PostgreSQL BIGINT range
                        if (
                            float_val < -9223372036854775808
                            or float_val > 9223372036854775807
                        ):
                            return None
                        return int(float_val)
                    else:  # DOUBLE PRECISION
                        return float_val
                except:
                    return None

            # Clean the value
            cleaned_val = re.sub(r"[^\d.\-+eE]", "", val_str)
            if cleaned_val in ("", ".", "-", "+"):
                return None

            # Validate numeric format
            if (
                cleaned_val.count(".") > 1
                or cleaned_val.count("-") > 1
                or cleaned_val.count("+") > 1
                or re.search(r"\d,\d,\d", val_str)
                or "/" in val_str
                or "+" in val_str[1:]
                or (target_type == "BIGINT" and re.search(r"[a-df-zA-DF-Z]", val_str))
            ):
                return None

            try:
                if target_type == "BIGINT":
                    float_val = float(cleaned_val)
                    if pd.isna(float_val) or float_val in (float("inf"), float("-inf")):
                        return None
                    # Check PostgreSQL BIGINT range
                    if (
                        float_val < -9223372036854775808
                        or float_val > 9223372036854775807
                    ):
                        return None
                    return int(float_val)
                else:  # DOUBLE PRECISION
                    return float(cleaned_val)
            except:
                return None

        else:
            # TEXT or fallback
            return str(value)


class ExcelUtils:
    """Utilities for Excel file cleaning."""

    @staticmethod
    async def clean_excel_pd(excel_file: BytesIO) -> dict[str, pd.DataFrame]:
        """
        This function cleans all sheets in an Excel file with pandas by:
        - forward-filling merged cells
        - removing empty rows and columns
        - filling NaN values with empty strings

        Returns a dictionary of dataframes with sheet names as keys.
        """
        wb = openpyxl.load_workbook(excel_file)
        tables = {}

        async def process_sheet(sheet_name: str):
            sheet = wb[sheet_name]

            # Create a dictionary to store merged cell values
            merged_cells_map = {}

            # Identify merged cells and store only the top-left value
            for merged_range in sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                top_left_value = sheet.cell(
                    row=min_row, column=min_col
                ).value  # Get top-left cell value

                # Store top-left value for all merged cells (but only modify in Pandas)
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        merged_cells_map[(row, col)] = top_left_value

            # Convert worksheet data into Pandas DataFrame
            df = pd.DataFrame(sheet.values)

            # Use Pandas to forward-fill merged cell values
            for (row, col), value in merged_cells_map.items():
                if pd.isna(
                    df.iloc[row - 1, col - 1]
                ):  # Adjust index for Pandas (0-based)
                    df.iloc[row - 1, col - 1] = value

            # Remove rows that are empty
            df.dropna(inplace=True, how="all")

            # Drop columns where all values are NaN
            df = df.dropna(axis=1, how="all")

            # Fill NaN values with empty strings for better readability
            df = df.fillna("")

            # Clean sheet name
            table_name = clean_table_name(sheet_name, existing=tables.keys())

            LOGGER.info(
                f"Sheet {sheet_name} after dropping NaN rows/columns: {df.shape[0]} rows, {df.shape[1]} columns"
            )

            return table_name, df

        async def main():
            with concurrent.futures.ThreadPoolExecutor() as executor:
                loop = asyncio.get_event_loop()
                tasks = [
                    loop.run_in_executor(
                        executor,
                        lambda sheet_name=sheet_name: asyncio.run(
                            process_sheet(sheet_name)
                        ),
                    )
                    for sheet_name in wb.sheetnames
                ]
                results = await asyncio.gather(*tasks)
                for table_name, df in results:
                    tables[table_name] = df

        await main()

        return tables

    @staticmethod
    async def is_excel_dirty(table_name: str, df: pd.DataFrame) -> bool:
        """
        Checks if an Excel dataframe needs additional cleaning with OpenAI.
        Returns True if the dataframe appears to need cleaning.
        
        Criteria for "dirty" Excel files:
        1. Headers/titles at the top that aren't part of the data
        2. Footnotes or notes at the bottom
        3. Rows where all values are the same (potential section headers)
        4. Rows with aggregate statistics (like "Total", "Sum", etc.)
        5. Wide format tables that should be transformed to long format
        """
        if df.empty or len(df) < 3:
            return False
            
        try:
            # Check for potential header/title rows at the top
            # Headers often have merged cells shown as the same value repeated
            head_rows = df.head(3)
            has_repeated_headers = head_rows.apply(
                lambda row: len(set(row.astype(str))) < len(row) * 0.5, axis=1
            ).any()
            
            # Check for potential footer rows
            tail_rows = df.tail(3)
            # Check for repeated values in footer rows (similar to headers)
            has_repeated_footers = tail_rows.apply(
                lambda row: len(set(row.astype(str))) < len(row) * 0.5, axis=1
            ).any()
            has_footer_notes = has_repeated_footers
            for _, row in tail_rows.iterrows():
                # Look for rows with few distinct values or text indicators
                text_values = [str(x).lower() for x in row if isinstance(x, str)]
                if text_values and any(note in " ".join(text_values) 
                                      for note in ["note", "source", "*", "total", "sum"]):
                    has_footer_notes = True
                    break
            
            # Check for rows where all non-null values are the same (section headers)
            same_value_rows = df[df.apply(
                lambda row: len(set(row.dropna())) == 1 and len(row.dropna()) > 1, axis=1
            )]
            has_section_headers = not same_value_rows.empty
            
            # Check for aggregate statistics rows (containing "total", "sum", etc.)
            has_aggregate_rows = False
            for _, row in df.iterrows():
                row_str = " ".join([str(x).lower() for x in row])
                if any(agg in row_str for agg in ["total", "sum", "subtotal", "average", "mean"]):
                    has_aggregate_rows = True
                    break
            
            # Check if it's in wide format (many columns with similar naming patterns)
            # Wide format often has repeated column name patterns
            col_names = [str(col).lower() for col in df.columns]
            repeated_patterns = []
            for i in range(len(col_names)):
                for j in range(i+1, len(col_names)):
                    # Check if columns follow patterns like "X 2020", "X 2021" or "Q1 X", "Q2 X"
                    pattern = re.findall(r'[a-z]+', col_names[i])
                    if pattern and any(p in col_names[j] for p in pattern):
                        repeated_patterns.append((col_names[i], col_names[j]))
            has_wide_format = len(repeated_patterns) > len(df.columns) * 0.3
            
            # Return True if any of the criteria are met
            is_dirty = has_repeated_headers or has_footer_notes or has_section_headers or has_aggregate_rows or has_wide_format
            
            if is_dirty:
                LOGGER.info(f"Excel sheet {table_name} requires further cleaning with OpenAI. Reasons: " +
                           f"repeated headers: {has_repeated_headers}, " +
                           f"footer notes: {has_footer_notes}, " +
                           f"section headers: {has_section_headers}, " +
                           f"aggregate rows: {has_aggregate_rows}, " +
                           f"wide format: {has_wide_format}")
            else:
                LOGGER.info(f"Excel sheet {table_name} is clean, skipping OpenAI cleaning")
            
            return is_dirty
            
        except Exception as e:
            LOGGER.error(f"Error checking if Excel is dirty: {e}")
            # If we encounter an error during checking, default to cleaning
            return True
    
    @staticmethod
    async def clean_excel_openai(table_name: str, df: pd.DataFrame) -> pd.DataFrame:
        """
        Further cleans a dataframe (from an Excel sheet) using OpenAI's Code Interpreter. Dynamically generates and executes code to remove columns and rows that do not contribute to the data (e.g. headers and footnotes). Also if necessary, changes the dataframe from wide to long format that's suitable for PostgreSQL.
        """
        # Check if the dataframe actually needs cleaning
        needs_cleaning = await ExcelUtils.is_excel_dirty(table_name, df)
        if not needs_cleaning:
            return df
            
        # Save CSV file in docker for upload to OpenAI
        file_path = f"./{table_name}.csv"
        df.to_csv(file_path, index=False)
        client = AsyncOpenAI()

        # Upload file using the File API
        try:
            csv_file = await client.files.create(
                file=open(file_path, "rb"), purpose="assistants"
            )
            LOGGER.info(
                f"Uploaded {table_name}.csv file to OpenAI for cleaning: {csv_file.id} "
            )
        except Exception as e:
            LOGGER.error(f"Failed to upload {table_name}.csv file to OpenAI: {e}")
            return df
        finally:
            # Delete file in docker
            os.remove(file_path)

        # Set up instructions and prompt
        instructions = "You are an expert in cleaning and transforming CSV files. Write and run code to execute transformations on CSV files."
        prompt = f"""Generate and execute a python script to clean and transform the provided CSV file that's been parsed from an Excel file.

    The script should perform the following tasks:
    0. Load the csv file as a dataframe
    1. Remove column indexes.
    2. Remove rows with titles or other plain text cells (e.g footnotes) that do not constitute the data. 
        - Inspect head and tail of dataframe.
        - Also inspect rows where all non-null values are the same with `df[df.apply(lambda row: len(set(row.dropna())) == 1, axis=1)]` to see if they are relevant data.
    3. Remove rows with aggregate statistics (i.e. Inspect rows with the word "total",” case-insensitively.)
    4. If table is in a wide format, change it to a long format so that it's suitable for a PostgreSQL database. Ensure no data is lost and that all columns are accounted for in the transformation. 
    5. Define meaningful column names for new columns.
    6. Generate a new CSV file with the cleaned and transformed data.

    Work with all the information you have. DO NOT ask further questions. Continue until the task is complete.
    """

        # Set up assistant, thread, message
        model = "gpt-4o"  # o3-mini currently doesn't support code interpreter
        assistant = await client.beta.assistants.create(
            instructions=instructions,
            model=model,
            tools=[{"type": "code_interpreter"}],
            tool_resources={"code_interpreter": {"file_ids": [csv_file.id]}},
        )
        thread = await client.beta.threads.create()

        message = await client.beta.threads.messages.create(
            thread_id=thread.id, role="user", content=prompt
        )

        # Run the code
        try:
            LOGGER.info(f"Executing excel cleaning run on {table_name}")
            run = await client.beta.threads.runs.create_and_poll(
                thread_id=thread.id,
                assistant_id=assistant.id,
                instructions=instructions,
            )
        except Exception as e:
            LOGGER.error(f"Failed to create and poll excel cleaning run: {e}")
            return df

        # Keep checking status of run
        if run.status == "completed":
            LOGGER.info(f"Excel cleaning run on {table_name} completed")
            messages = await client.beta.threads.messages.list(thread_id=thread.id)
        else:
            LOGGER.error(
                f"Excel cleaning run on {table_name} did not complete successfully. Run status: {run.status}"
            )
            return df

        # Extract file to download
        file_to_download = None
        for m in messages.data:
            for content_block in m.content:
                if hasattr(content_block, "text") and hasattr(
                    content_block.text, "annotations"
                ):
                    for annotation in content_block.text.annotations:
                        if hasattr(annotation, "file_path") and hasattr(
                            annotation.file_path, "file_id"
                        ):
                            file_to_download = annotation.file_path.file_id
                            break

        # Download the file and convert to dataframe
        if file_to_download is not None:
            try:
                file_data = await client.files.content(file_to_download)
            except Exception as e:
                LOGGER.error(f"Failed to download {file_to_download}: {e}")
                return df

            try:
                file_data_bytes = file_data.read()
                df = pd.read_csv(BytesIO(file_data_bytes))
                LOGGER.info(f"Downloaded {file_to_download}")
            except Exception as e:
                LOGGER.error(f"Failed to read {file_to_download} as CSV: {e}")
                return df

            # Delete file in client
            await client.files.delete(file_to_download)
            LOGGER.info(f"Deleted {file_to_download} in client")
        else:
            LOGGER.info(f"No file to download.")

        # Calculate run cost
        LOGGER.info(f"Run usage: {run.usage}")
        output_tokens = run.usage.completion_tokens
        cached_input_tokens = run.usage.prompt_token_details.get("cached_tokens", 0)
        input_tokens = run.usage.prompt_tokens - cached_input_tokens

        cost = input_tokens / 1000 * LLM_COSTS_PER_TOKEN[model]["input_cost_per1k"]
        cost += output_tokens / 1000 * LLM_COSTS_PER_TOKEN[model]["output_cost_per1k"]
        cost += (
            cached_input_tokens
            / 1000
            * LLM_COSTS_PER_TOKEN[model]["cached_input_cost_per1k"]
        )
        cost *= 100
        LOGGER.info(f"Run cost in cents: {cost}")
        return df


class DbUtils:
    """Utilities for database operations."""

    @staticmethod
    def create_table_sql(table_name: str, columns: dict[str, str]):
        """
        Build a CREATE TABLE statement.

        Args:
            table_name: Name of the table to create
            columns: Dictionary mapping column names to data types

        Returns:
            SQL statement for creating the table
        """
        cols = []
        for col_name, col_type in columns.items():
            safe_col_name = NameUtils.sanitize_column_name(col_name)
            cols.append(f'"{safe_col_name}" {col_type}')

        cols_str = ", ".join(cols)
        return f'CREATE TABLE "{table_name}" ({cols_str});'

    @staticmethod
    async def export_df_to_postgres(
        df: pd.DataFrame,
        table_name: str,
        db_connection_string: str,
        chunksize: int = 5000,
    ):
        """
        Export a pandas DataFrame to PostgreSQL.

        Args:
            df: DataFrame to export
            table_name: Name of the target table
            db_connection_string: Database connection string
            chunksize: Number of rows to insert at once

        Returns:
            Dictionary with success status and inferred types
        """
        # Make a copy and handle NaN values
        df = df.copy().fillna(value="")

        # Store original column names for type inference
        original_cols = list(df.columns)

        # Sanitize column names and handle duplicates
        safe_col_list = []
        seen_names = set()

        for col in df.columns:
            safe_name = NameUtils.sanitize_column_name(col)

            # Handle duplicate sanitized names
            if safe_name in seen_names:
                counter = 1
                while f"{safe_name}_{counter}" in seen_names:
                    counter += 1
                safe_name = f"{safe_name}_{counter}"

            safe_col_list.append(safe_name)
            seen_names.add(safe_name)

        # Create mapping between sanitized and original names
        col_name_mapping = dict(zip(safe_col_list, original_cols))

        # Update dataframe with sanitized column names
        df.columns = safe_col_list

        # Create a SQLAlchemy engine
        engine = create_async_engine(db_connection_string)

        # Infer data types using original column names
        inferred_types = {}
        for col in df.columns:
            original_name = col_name_mapping.get(col, col)
            try:
                inferred_types[col] = TypeUtils.guess_column_type(
                    df[col], column_name=original_name
                )
            except Exception as e:
                raise Exception(
                    f"Failed to infer type for column {original_name} in table {table_name}: {e}"
                )

        LOGGER.info(inferred_types)

        # Convert values to appropriate PostgreSQL types
        converted_df = df.copy()
        for col in df.columns:
            try:
                converted_df[col] = df[col].map(
                    lambda value: TypeUtils.convert_values_to_postgres_type(
                        value, target_type=inferred_types[col]
                    )
                )
            except Exception as e:
                raise Exception(
                    f"Failed to convert values to PostgreSQL type for column {col} in table {table_name}: {e}"
                )

        # Create table in PostgreSQL
        try:
            create_stmt = DbUtils.create_table_sql(table_name, inferred_types)
        except Exception as e:
            raise Exception(
                f"Failed to create CREATE TABLE statements for {table_name}: {e}"
            )
        async with engine.begin() as conn:
            await conn.execute(text(f'DROP TABLE IF EXISTS "{table_name}";'))
            await conn.execute(text(create_stmt))

        # Prepare and execute INSERT statements
        insert_cols = ", ".join(f'"{c}"' for c in safe_col_list)
        placeholders = ", ".join([f":{c}" for c in safe_col_list])
        insert_sql = (
            f'INSERT INTO "{table_name}" ({insert_cols}) VALUES ({placeholders})'
        )

        async with engine.begin() as conn:
            rows_to_insert = []
            for idx, row in enumerate(
                converted_df.replace({np.nan: None}).to_dict("records")
            ):
                rows_to_insert.append(row)

                # Batch insert when chunk size reached or at the end
                if len(rows_to_insert) == chunksize or idx == len(converted_df) - 1:
                    await conn.execute(text(insert_sql), rows_to_insert)
                    rows_to_insert = []

        print(f"Successfully imported {len(df)} rows into table '{table_name}'.")
        return {"success": True, "inferred_types": inferred_types}


# Legacy function aliases for backward compatibility
def clean_table_name(table_name: str, existing=None):
    """Legacy wrapper for NameUtils.clean_table_name"""
    return NameUtils.clean_table_name(table_name, existing or [])


def sanitize_column_name(col_name: str):
    """Legacy wrapper for NameUtils.sanitize_column_name"""
    return NameUtils.sanitize_column_name(col_name)


def is_date_column_name(col_name):
    """Legacy wrapper for DateTimeUtils.is_date_column_name"""
    return DateTimeUtils.is_date_column_name(col_name)


def is_time_column_name(col_name):
    """Legacy wrapper for DateTimeUtils.is_time_column_name"""
    return DateTimeUtils.is_time_column_name(col_name)


def can_parse_date(val):
    """Legacy wrapper for DateTimeUtils.can_parse_date"""
    return DateTimeUtils.can_parse_date(val)


def can_parse_time(val):
    """Legacy wrapper for DateTimeUtils.can_parse_time"""
    return DateTimeUtils.can_parse_time(val)


def to_float_if_possible(val):
    """Legacy wrapper for TypeUtils.to_float_if_possible"""
    return TypeUtils.to_float_if_possible(val)


def guess_column_type(series, column_name=None, sample_size=50):
    """Legacy wrapper for TypeUtils.guess_column_type"""
    return TypeUtils.guess_column_type(series, column_name, sample_size)


def convert_values_to_postgres_type(value, target_type: str):
    """Legacy wrapper for TypeUtils.convert_values_to_postgres_type"""
    return TypeUtils.convert_values_to_postgres_type(value, target_type)


def create_table_sql(table_name: str, columns: dict[str, str]):
    """Legacy wrapper for DbUtils.create_table_sql"""
    return DbUtils.create_table_sql(table_name, columns)


async def export_df_to_postgres(
    df: pd.DataFrame, table_name: str, db_connection_string: str, chunksize: int = 5000
):
    """Legacy wrapper for DbUtils.export_df_to_postgres"""
    return await DbUtils.export_df_to_postgres(
        df, table_name, db_connection_string, chunksize
    )
